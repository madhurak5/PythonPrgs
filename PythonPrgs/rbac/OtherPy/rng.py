import random
from pathlib import Path
import openpyxl as xl
from openpyxl.chart import BarChart, Reference



members = ["John", "Mary", "Bob", "Tim"]
for i in range(3):
    print(random.random())  # generates a random val between 0 and 1
    print(random.randint(10, 30))
leader = random.choice(members)
print(f"Leader : {leader}")


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        sec = random.randint(1, 6)
        return first, sec


d = Dice()
a = d.roll()
print(a)
"""
path = Path("eComm1")
print(path.exists())
path = Path("Email")
print(" directory created", path.name, path.mkdir())

print(f"Deleting {path.name} {path.rmdir()}")



path = Path()
for file in (path.glob('*')):
    print(path)

"""
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1,1)
    print("Cell value : ",cell.value)
    total_rows = sheet.max_row
    for row in range(2, total_rows+1):
        print(row)
        cell1 = sheet.cell(row, 3)
        price = cell1.value
        print(price)
        try:
            corrected_price = price * 0.9
        except: FloatingPointError
        print("Corrected price : ", corrected_price)
        corrected_cell = sheet.cell(row, 4)
        corrected_cell.value = corrected_price

    values = Reference(sheet,
               min_row=2,
               max_row=total_rows,
               min_col=4,
               max_col=4
            )
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')
    wb.save(filename)


fname = input("Enter the filename : ")
process_workbook(fname)